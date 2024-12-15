import requests
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from io import BytesIO

# Step 1: Download the Google Sheets file as an Excel file (.xlsx)
url = 'https://docs.google.com/spreadsheets/d/1kW6SZYKSCZc5XqaWW0oTFYNtN41WIWkCp4p4Uku7aRc/export?format=xlsx'

local_filename = 'downloaded_file.xlsx'

# Download the file and save it locally
response = requests.get(url)
with open(local_filename, 'wb') as file:
    file.write(response.content)

# Step 2: Load the downloaded Excel file using openpyxl
wb = load_workbook(local_filename)

# Select the active sheet or specify a particular sheet by name
ws = wb.active  # or ws = wb['SheetName']

# Step 3: Locate the "Pictures" column by header name
header_row = ws[1]  # The first row (header)
pictures_col_idx = 4

for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
    # Extract the name from the first column (assuming the name is in the first cell of the row)
    first_name, last_name, position, email, description = [el.value for el in row][:5]
    image_name = first_name.strip()[0]+last_name.strip().replace(" ", "_").upper()
    image_path = f"../assets/images/people/{image_name}.jpg"
    if not image_name:
        print(f"Skipping row {row_idx}, no name found.")
        continue

    # Map the image to the row without printing the image object
    try:
        if len(ws._images) >= row_idx - 1:  # Assumes one image per row
            image = ws._images[row_idx - 2]  # Match image by row index (adjusted for zero-index)
            img_stream = BytesIO(image._data())  # Access the raw image data
            pil_image = PILImage.open(img_stream)

            # Save the image using the name from the first column
            pil_image.save(image_path, format='JPEG')

            print(f"Image for '{image_name}' saved as {image_path}")
        else:
            print(f"No image found for row {row_idx}")
    except Exception as e:
        print(f"Error processing image for '{image_name}' in row {row_idx}: {e}")

    with open(f"../_members/{image_name}.md", "w") as out:
        out.write(
f"""---
layout: people
name: {first_name} {last_name}
type: {position}
position: {position}
image: {image_path}
email: {email}
---
{description}
""")
